import json
import requests
from google.auth.transport.requests import Request
from google.auth import default
import os
import openpyxl
from openpyxl.styles import PatternFill

class BigQueryTagManager:
    def __init__(self, project_id, dataset_id, tag_key_id, tag_value_id):
        self.project_id = project_id
        self.dataset_id = dataset_id
        self.tag_key_id = tag_key_id
        self.tag_value_id = tag_value_id

        # tags in the required format Tag key is expected to be in the namespaced format, for example "123456789012/environment" where 123456789012 is the ID of the parent organization or project resource for this tag key. Tag value is expected to be the short name, for example "Production".
        self.resource_tags = {f'{self.project_id}/{self.tag_key_id}': self.tag_value_id}
        self.credentials, _ = default(scopes=["https://www.googleapis.com/auth/cloud-platform"])
        auth_request = Request()
        self.credentials.refresh(auth_request)

        self.headers = {
            "Authorization": f"Bearer {self.credentials.token}",
            "Content-Type": "application/json"
        }

    def apply_tags_to_dataset(self):
        """Apply tags to the dataset."""
        dataset_url = f"https://bigquery.googleapis.com/bigquery/v2/projects/{self.project_id}/datasets/{self.dataset_id}"
        dataset_body = {"resourceTags": self.resource_tags}

        response = requests.put(dataset_url, headers=self.headers, data=json.dumps(dataset_body))
        if response.status_code == 200:
            print(f"Tags updated successfully on dataset {self.dataset_id}: {response.json()}")
        else:
            print(f"Failed to update tags on dataset {self.dataset_id}: {response.status_code} {response.text}")

    def list_tables(self):
        """List all tables in the dataset, handling pagination if there are more than 50 tables."""
        tables_url = f"https://bigquery.googleapis.com/bigquery/v2/projects/{self.project_id}/datasets/{self.dataset_id}/tables"
        tables = []
        page_token = None

        while True:
            # Add pageToken to the request if it exists (for pagination)
            params = {'pageToken': page_token} if page_token else {}
            response = requests.get(tables_url, headers=self.headers, params=params)

            if response.status_code == 200:
                response_data = response.json()
                tables.extend([table["tableReference"]["tableId"] for table in response_data.get("tables", [])])
                
                # Check if there are more tables (nextPageToken is present)
                page_token = response_data.get("nextPageToken")
                if not page_token:
                    break  # No more pages, exit the loop
            else:
                print(f"Failed to list tables: {response.status_code} {response.text}")
                break

        return tables

    def fetch_table_metadata(self, table_id):
        """Fetch the metadata of a specific table."""
        table_url = f"https://bigquery.googleapis.com/bigquery/v2/projects/{self.project_id}/datasets/{self.dataset_id}/tables/{table_id}"
        response = requests.get(table_url, headers=self.headers)

        if response.status_code == 200:
            return response.json()
        else:
            print(f"Failed to fetch metadata for table {table_id}: {response.status_code} {response.text}")
            return None

    def compare_and_generate_report(self, tables, report_filename):
        """Generate a single Excel report with detailed field-level comparison."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Metadata Comparison"
        ws.append(["Table Name", "Field", "Before", "After", "Comparison"])

        change_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        no_change_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

        current_row = 2

        for table_id in tables:
            before_metadata = self.fetch_table_metadata(table_id)
            if not before_metadata:
                continue

            self.apply_tags_to_tables([table_id])
            after_metadata = self.fetch_table_metadata(table_id)

            if not after_metadata:
                continue

            all_fields = set(before_metadata.keys()).union(after_metadata.keys())
            for field in all_fields:
                before_value = before_metadata.get(field, "N/A")
                after_value = after_metadata.get(field, "N/A")
                comparison = "No Change" if before_value == after_value else "Changed"
                
                if isinstance(before_value, dict):
                    before_value = str(before_value)  
                if isinstance(after_value, dict):
                    after_value = str(after_value)

                row = [table_id, field, before_value, after_value, comparison]
                ws.append(row)

                for cell in ws[current_row][2:]:
                    if comparison == "Changed":
                        cell.fill = change_fill
                    else:
                        cell.fill = no_change_fill

                current_row += 1

        wb.save(report_filename)
        print(f"Report generated: {report_filename}")


    def apply_tags_to_tables(self, tables):
        """Apply tags to each table in the dataset."""
        for table_id in tables:
            table_url = f"https://bigquery.googleapis.com/bigquery/v2/projects/{self.project_id}/datasets/{self.dataset_id}/tables/{table_id}"
            response = requests.get(table_url, headers=self.headers)

            if response.status_code == 200:
                table_metadata = response.json()
                table_metadata["resourceTags"] = self.resource_tags

                patch_response = requests.patch(table_url, headers=self.headers, data=json.dumps(table_metadata))
                if patch_response.status_code == 200:
                    print(f"Tags updated successfully on table {table_id}: {patch_response.json()}")
                else:
                    print(f"Failed to update tags on table {table_id}: {patch_response.status_code} {patch_response.text}")
            else:
                print(f"Failed to fetch metadata for table {table_id}: {response.status_code} {response.text}")


if __name__ == "__main__":
    # Read values from environment variables or hardcode for testing
    project_id = os.getenv("PROJECT_ID", "your_project_id")
    dataset_id = os.getenv("DATASET_ID", "your_dataset_id")
    tag_key_id = os.getenv("TAG_KEY_ID", "your_tag_key_id")
    tag_value_id = os.getenv("TAG_VALUE_ID", "your_tag_value_id")

    tag_manager = BigQueryTagManager(project_id, dataset_id, tag_key_id, tag_value_id)

    tables = tag_manager.list_tables()

    if tables:
        print(f"Found {len(tables)} tables to apply tags to.")
        tag_manager.compare_and_generate_report(tables, "metadata_report_new.xlsx")
    else:
        print("No tables found in the dataset.")