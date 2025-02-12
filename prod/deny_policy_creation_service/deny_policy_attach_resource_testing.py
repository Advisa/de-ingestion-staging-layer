import json
import requests
import openpyxl
from openpyxl.styles import PatternFill
from google.auth.transport.requests import Request
from google.auth import default

def get_bearer_token():
    """Fetches the Bearer token using Google Cloud credentials."""
    # Load default credentials from the environment or local configuration
    credentials, _ = default(scopes=["https://www.googleapis.com/auth/cloud-platform"])
    
    # Refresh the token
    credentials.refresh(Request())
    
    return credentials.token

def list_tables(project_id, dataset_id, headers):
    """List all tables in the dataset."""
    base_url = "https://bigquery.googleapis.com/bigquery/v2"
    url = f"{base_url}/projects/{project_id}/datasets/{dataset_id}/tables"
    tables = []
    page_token = None

    while True:
        params = {'pageToken': page_token} if page_token else {}
        response = requests.get(url, headers=headers, params=params)

        if response.status_code == 200:
            response_data = response.json()
            tables.extend([table["tableReference"]["tableId"] for table in response_data.get("tables", [])])

            page_token = response_data.get("nextPageToken")
            if not page_token:
                break
        else:
            print(f"Error fetching tables: {response.status_code}, {response.text}")
            break
    return tables

def fetch_table_metadata(project_id, dataset_id, table_id, headers):
    """Fetch metadata of a table."""
    base_url = "https://bigquery.googleapis.com/bigquery/v2"
    table_url = f"{base_url}/projects/{project_id}/datasets/{dataset_id}/tables/{table_id}"
    response = requests.get(table_url, headers=headers)

    if response.status_code == 200:
        return response.json()
    else:
        print(f"Error fetching metadata for table {table_id}: {response.status_code}, {response.text}")
        return None

def generate_report(project_id, datasets, headers, report_filename):
    """Generate an Excel report for specified datasets and tables with metadata."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Table Metadata Report"
    
    # Add headers for the report
    ws.append(["Dataset", "Table", "Field", "Value", "Resource Tag Match"])

    # Highlight cells
    mismatch_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    match_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    expected_resource_tag = {'data-domain-data-warehouse/gdpr_complaince_tag_prod': 'gdpr_complaince_5year_prod'}

    current_row = 2
    for dataset_id in datasets:
        print(dataset_id)
        tables = list_tables(project_id, dataset_id, headers)
        
        for table_id in tables:
            print(table_id)
            table_metadata = fetch_table_metadata(project_id, dataset_id, table_id, headers)
            if not table_metadata:
                continue

            resource_tags = table_metadata.get("resourceTags", {})
            resource_tag_match = "Yes" if resource_tags == expected_resource_tag else "No"

            # Add dataset and table information
            ws.append([dataset_id, table_id, "resourceTags", str(resource_tags), resource_tag_match])

            # Check if the tags match the expected ones and color accordingly
            if resource_tag_match == "No":
                for cell in ws[current_row][4:]:
                    cell.fill = mismatch_fill
            else:
                for cell in ws[current_row][4:]:
                    cell.fill = match_fill

            # Check all metadata fields and add them to the report
            for field, value in table_metadata.items():
                if field != "resourceTags":
                    row = [dataset_id, table_id, field, str(value), ""]
                    ws.append(row)
                    current_row += 1
        
            current_row += 1

    # Save the workbook to a file
    wb.save(report_filename)
    print(f"Report generated: {report_filename}")


# Usage
project_id = "data-domain-data-warehouse"  # Replace with your actual project ID
datasets = ["sambla_group_data_stream","sambla_group_data_stream_fi","sambla_group_data_stream_no","sambla_new_mongodb"]  # List your datasets here
bearer_token = get_bearer_token()  # Fetch the Bearer token using Google credentials

# Authorization headers with the dynamic Bearer token
headers = {
    "Authorization": f"Bearer {bearer_token}",
    "Content-Type": "application/json"
}

# Generate the report with table metadata and resource tags for the specified datasets
generate_report(project_id, datasets, headers, "table_metadata_report_selected_datasets.xlsx")
